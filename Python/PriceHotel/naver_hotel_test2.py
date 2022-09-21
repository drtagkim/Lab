#author: taekyung kim
#test update: 2022-09-20

from playwright.sync_api import sync_playwright,expect,TimeoutError
from os.path import exists
import pandas as pd
from jsonpath_ng.ext import parse #extension parser for JSONPath implementation
import datetime
from datetime import timedelta
from openpyxl import load_workbook
import yaml
import sys

yaml_input="input.yaml"

def pick_date_from_today(delta=0):
    t=datetime.date.today()
    t1=t+timedelta(days=delta) #하루를 더한다.
    return t1.strftime("%Y-%m-%d")

def combine_list(x):
    if isinstance(x,list):
        y=",".join(x)
    else:
        y=x
    return y

def filter_data(res):
    p1=parse("$..rates[*].hcHotelId")
    p2=parse("$..rates[*].otaCode")
    p3=parse("$..rates[*].availableRoomCnt")
    p4=parse("$..rates[*].totalRate")
    p5=parse("$..rates[*].promotionTotalRate")
    p6=parse("$..rates[*].promotionBenefitRate")
    p7=parse("$..rates[*].promotionBenefitType")
    p8=parse("$..rates[*].promotionBenefitValue")
    p9=parse("$..rates[*].roomCategories")
    output=pd.DataFrame({
        "hcHotelId":[i.value for i in p1.find(res)],
        "otaCode":[i.value for i in p2.find(res)],
        "availableRoomCnt":[i.value for i in p3.find(res)],
        "totalRate":[i.value for i in p4.find(res)],
        "promotionTotalRate":[i.value for i in p5.find(res)],
        "promotionBenefitRate":[i.value for i in p6.find(res)],
        "promotionBenefitType":[i.value for i in p7.find(res)],
        "promotionBenefitValue":[i.value for i in p8.find(res)],
        "roomCategories":[combine_list(i.value) for i in p9.find(res)],
    })
    return output

def write_data_to_excel(df,file_name):
    book = load_workbook(file_name)
    writer = pd.ExcelWriter(file_name, engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    for sheetname in writer.sheets:
        df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)
    writer.save()

class OutputData:
    def __init__(self):
        self.df=None
        self.checked=False
def run_main(fout_file,url,res_wait=500,time_span=0):
    #글래드 여의도(test)
    #url="https://hotels.naver.com/item?hotelFileName=hotel%3AGLAD_YEOUIDO"
    url+="&adultCnt=2"
    url+=f"&checkIn={pick_date_from_today(time_span)}"
    url+=f"&checkOut={pick_date_from_today(time_span+1)}"
    #print(url)
    df=OutputData()
    tracking=0
    with sync_playwright() as p:
        def handle_response(response):
            if 'https://hotel-api.naver.com/graphql' in response.url:
                try:
                    response_data=response.json()
                    p=parse("$..rates") #discover top rates
                    m=p.find(response_data)
                    if len(m)>0:
                        df.df=filter_data(response_data)
                        df.checked=True
                except:
                    pass
        def wait_event_complete(response):
            if df.checked:
                return True
            else:
                return False
        b=p.chromium.launch_persistent_context(user_data_dir='./temp',headless=True)
        #b=p.webkit.launch_persistent_context(user_data_dir='./temp',headless=True)
        p=b.new_page()
        p.on('response',handle_response)
        while True:
            tracking+=1
            try:
                p.goto(url,wait_until='commit',timeout=1000*30) #30 seconds
                break
            except TimeoutError:
                assert tracking<10,"Too many trials"
                p.wait_for_timeout(5000)
                continue
        #가격비교탭 클릭 - more
        #btn1=p.locator('div.common_Tab__CS5JL > div > a:nth-child(2)')
        #expect(btn1).not_to_be_empty()
        #btn1.click()
        #btn2=p.locator('button.Rates_Date__LaozD.Rates_btn_checkin__vo3SG')
        #expect(btn2).not_to_be_empty()
        #btn2.click()
        #p.wait_for_event('response',wait_event_complete,timeout=1000*30)
        p.wait_for_timeout(res_wait)
        p.context.close()
        b.close()
    if df.df is not None:
        df.df['url']=url
        df.df['today']=pick_date_from_today()
        df.df['pickDate']=pick_date_from_today(time_span)
        if not exists(fout_file):
            df.df.to_excel(fout_file,index=False)
        else:
            write_data_to_excel(df.df,fout_file)
        #print("result out.")
if __name__=="__main__":
    with open(yaml_input) as f:
        settings=yaml.load(f,Loader=yaml.FullLoader)
    assert "setting" in settings,"Setting should be!"
    fout_file=settings['setting']['foutfile']
    url=settings['setting']['url']
    res_wait=settings['setting']['res_wait']
    for i in range(settings['setting']['days']):
        print(f"{fout_file}...{i+1}")
        run_main(fout_file,url,res_wait,i)
    print("OK.")
